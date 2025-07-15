import os
import datetime
from openpyxl import Workbook, load_workbook

def print_header():
    print("\n" + "="*60)
    print("✅ All rights reserved to Mohamed Elyes Lazher ✅")
    print("="*60 + "\n")

def get_excel_filename():
    today = datetime.date.today().strftime("%Y-%m-%d")
    return f"scan_log_{today}.xlsx"

def initialize_excel(filename):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Scans"
        ws.append(["Timestamp", "Box No", "Scan Type", "Scanned Qty", "Good Qty", "NG Qty", "SNs"])
        wb.save(filename)

def log_to_excel(filename, box_no, scan_type, scanned_qty, good_qty, ng_qty, sn_list):
    wb = load_workbook(filename)
    ws = wb["Scans"]
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sn_text = ", ".join(sn_list)
    ws.append([now, box_no, scan_type, scanned_qty, good_qty, ng_qty, sn_text])
    wb.save(filename)

def main():
    print_header()
    excel_file = get_excel_filename()
    initialize_excel(excel_file)

    while True:
        input("\n📦 Scan to OPEN the box (press Enter after scan): ")
        box_no = input("➡️  Box Number: ").strip()
        print(f"✅ Box {box_no} opened. Begin scanning up to 36 unique pieces.\n")

        scanned_items = set()

        while True:
            sn = input(f"🔄 Scan piece {len(scanned_items)+1} (or re-scan box number to CLOSE): ").strip()
            if sn == "":
                print("⚠️  Empty input, please scan again.")
                continue
            if sn == box_no:
                print(f"📦 Box {box_no} closed by re-scanning box number.")
                break
            elif sn in scanned_items:
                print(f"⚠️  This SN has already been scanned! Duplicate ignored.")
                continue
            else:
                scanned_items.add(sn)
                print(f"✅ Piece {len(scanned_items)} scanned successfully.")

            if len(scanned_items) >= 36:
                print(f"\n✅ 36 pieces scanned. Closing box {box_no} automatically.")
                break

        good_qty = len(scanned_items)
        ng_qty = 36 - good_qty

        print("\n📊 Résumé du box:")
        print(f"✔️  Box No: {box_no}")
        print(f"✔️  Checked Qty: {good_qty}")
        print(f"✔️  Good Qty: {good_qty}")
        print(f"❌ NG Qty: {ng_qty}")

        log_to_excel(
            excel_file,
            box_no,
            "Completed",
            good_qty,
            good_qty,
            ng_qty,
            list(scanned_items)
        )

        print(f"💾 Données enregistrées dans {excel_file}\n")
        repeat = input("➡️  Voulez-vous scanner une autre boîte ? (y/n): ").lower()
        if repeat != "y":
            print("\n✅ Merci ! Fin du programme.")
            break

if __name__ == "__main__":
    main()
